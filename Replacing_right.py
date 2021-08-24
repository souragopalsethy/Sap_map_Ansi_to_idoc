import  openpyxl

wb_op_replacing=openpyxl.load_workbook("C:\\Users\\Souragopal\\Desktop\\py_excel 875\\ran_2.xlsx");

sheet_op_replacing=wb_op_replacing["850r41XB"]

outputSheet_name = "875r41uge"
outputFile_name = "Hershey's_"+outputSheet_name+" MappingSheet.xlsx"

sheet_op_replacing.cell(1, 1, "Gentran Mappingid : "+outputSheet_name+"\nTrading Partner code : ")


row_count=426
count=2

not_found_wb=openpyxl.load_workbook("C:\\Users\\Souragopal\\Desktop\\py_excel 875\\Program_base_Excels\\noFound.xlsx");
not_found_sheet=not_found_wb["right"]

for y in range(4, row_count):

    segment_data = sheet_op_replacing.cell(y, 8).value

    target_field_desc=sheet_op_replacing.cell(y, 9).value

    temp_tfd=str(target_field_desc)

    if("-" in temp_tfd):
        temp = str(target_field_desc).split("-")

        mapping_data_segment = temp[0]
        for z in range(4, row_count):
            mapping_data = sheet_op_replacing.cell(z, 6).value
            #if not segment_data :
            mapping_data_new = str(mapping_data).replace(str(segment_data), str(mapping_data_segment))
            sheet_op_replacing.cell(row=z, column=6, value=mapping_data_new)  # mapping

        sheet_op_replacing.cell(row=y, column=8, value=mapping_data_segment) #to set taregt field

    else:
        not_found_sheet.cell(row=count, column=1, value=y)
        not_found_sheet.cell(row=count, column=2, value=target_field_desc)
        count = count + 1



not_found_wb.save("not_found_seg.xlsx")

sheet_op_replacing.title = outputSheet_name
wb_op_replacing.save(outputFile_name)
