import  openpyxl

main_wb=openpyxl.load_workbook("C:\\\Users\\\\Souragopal\\\\Desktop\\\\py_excel 875\\\\raunakbabu875r41uge.xlsx");
main_sheet=main_wb["850r41XB"]

ref_wb=openpyxl.load_workbook("C:\\Users\\Souragopal\\Desktop\\py_excel 875\\Program_base_Excels\\ref_id.xlsx");
ref_sheet=ref_wb["data"]

not_found_wb=openpyxl.load_workbook("C:\\Users\\Souragopal\\Desktop\\py_excel 875\\Program_base_Excels\\noFound.xlsx");
not_found_sheet=not_found_wb["left"]

LAST_ROW=432
count=2
for z in range(4, LAST_ROW):
    main_ss=main_sheet.cell(z, 1).value

    main_sf = main_sheet.cell(z, 2).value

    main_ml = main_sheet.cell(z, 6).value

    flag=False

    for a in range(2, 88):
        ref=ref_sheet.cell(a, 1).value
        id = ref_sheet.cell(a, 2).value
        new_id="D_"+str(id)
        if str(main_ss) in str(ref):
            main_sheet.cell(row=z, column=2, value=new_id)  # seg ss
            for d in range(4, LAST_ROW):
                main_ml = main_sheet.cell(d, 6).value
                if str(main_sf)in str(main_ml):
                    temp= str(main_ml).replace(str(main_sf), str(new_id))
                    main_sheet.cell(row=d, column=6, value=temp)  # mapping
            flag=True
            print("works")
    if(flag==False):
        not_found_sheet.cell(row=count, column=1, value=z)
        not_found_sheet.cell(row=count, column=2, value=main_ss)
        count = count + 1





main_wb.save("ran_2.xlsx") #to crratye a ndw file
not_found_wb.save("not_found_seg.xlsx")