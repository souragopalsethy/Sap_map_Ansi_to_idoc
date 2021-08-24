import  openpyxl
wb=openpyxl.load_workbook("bada.xlsx");
# load persheetticular
S_NAME="875r41u"
sheet=wb[S_NAME]
FILE_NAME=S_NAME;


wb_op=openpyxl.load_workbook("C:\\Users\\Souragopal\\Desktop\\py_excel 875\\Ref_blank.xlsx");
# load persheetticular
sheet_op=wb_op["850r41XB"]

row_count=sheet_op.max_row
 # NED TO CHANGE THE LOOP CONDITION
for x in range(3, 431):
    #Target_side
    mapping = sheet.cell(x, 9).value
    sheet_op.cell(row=x + 1, column=6, value=mapping)  # mapping

    rec=sheet.cell(x, 12).value
    sheet_op.cell(row=x+1,column=7,value=rec) #targetsegment

    idoc_src = sheet.cell(x, 13).value
    sheet_op.cell(row=x + 1, column=8, value=idoc_src) #targetfiled

    idoc_name = sheet.cell(x, 14).value
    sheet_op.cell(row=x + 1, column=9, value=idoc_name)  # targetfileddescription

    Src_idoc_occurance = sheet.cell(x, 11).value
    sheet_op.cell(row=x + 1, column=10, value=Src_idoc_occurance)  # target_idoc_occurance

    #Src_idoc_length = sheet.cell(x, 18).value
    s_ansi_min = sheet.cell(x, 16).value
    s_ansi_max = sheet.cell(x, 17).value
    s_ansi_datatype = "Num" + "(" + str(s_ansi_min) + "-" + str(s_ansi_max) + ")"
    sheet_op.cell(row=x + 1, column=11, value=s_ansi_datatype)  # target_idoc_length


    #source side code

    ansi_seg=sheet.cell(x, 3).value
    sheet_op.cell(row=x + 1, column=1, value=ansi_seg)  # source_Ansi_Seg

    ansi_dest_field = sheet.cell(x, 4).value
    sheet_op.cell(row=x + 1, column=2, value=ansi_dest_field)  # ansi_source_field

    ansi_name = sheet.cell(x, 8).value
    sheet_op.cell(row=x + 1, column=3, value=ansi_name)  # ansi_source_desc

    ansi_occurance = sheet.cell(x, 2).value
    sheet_op.cell(row=x + 1, column=4, value=ansi_occurance)  # ansi_source_occurance

    ansi_min = sheet.cell(x, 6).value
    ansi_max = sheet.cell(x, 7).value
    ansi_datatype="Num"+"("+str(ansi_min)+"-"+str(ansi_max)+")"
    sheet_op.cell(row=x + 1, column=5, value=ansi_datatype)  # ansi_source_datatype



wb_op.save("sombit_1_"+FILE_NAME+".xlsx")




